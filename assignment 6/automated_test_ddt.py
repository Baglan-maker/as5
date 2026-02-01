
import openpyxl
from datetime import datetime
from typing import Dict, List, Tuple
import time
import sys


class TestExecutionTracker:

    def __init__(self):
        self.execution_start_time = datetime.now()
        self.test_outcomes = []
        self.total_assertions = 0
        self.passed_assertions = 0

    def record_outcome(self, test_id: str, status: str, details: str):
        """Records individual test result"""
        self.test_outcomes.append({
            'id': test_id,
            'status': status,
            'details': details,
            'timestamp': datetime.now()
        })
        self.total_assertions += 1
        if status == "PASSED":
            self.passed_assertions += 1

    def generate_summary_report(self) -> str:
        """Creates formatted summary of test execution"""
        duration = (datetime.now() - self.execution_start_time).total_seconds()
        pass_rate = (self.passed_assertions / self.total_assertions * 100) if self.total_assertions > 0 else 0

        summary = f"""
    Execution Duration: {duration:.2f} seconds                   
    Total Test Cases: {self.total_assertions}                    
    Passed: {self.passed_assertions}                             
    Failed: {self.total_assertions - self.passed_assertions}     
    Pass Rate: {pass_rate:.1f}%                                  
"""
        return summary


class ExcelDataProvider:
    """Handles reading test data from Excel file"""

    def __init__(self, filepath: str, sheet_name: str):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None

    def initialize_connection(self):
        """Opens Excel file and selects worksheet"""
        try:
            self.workbook = openpyxl.load_workbook(self.filepath)
            self.worksheet = self.workbook[self.sheet_name]
            print(f"✓ Successfully loaded Excel file: {self.filepath}")
            print(f"✓ Active sheet: {self.sheet_name}")
            return True
        except Exception as error:
            print(f"✗ Error loading Excel file: {error}")
            return False

    def extract_test_scenarios(self) -> List[Dict]:
        """Extracts all test scenarios from Excel sheet"""
        scenarios_collection = []

        # Read header row to get column mapping
        header_row = [cell.value for cell in self.worksheet[1]]

        # Process each data row (starting from row 2)
        for row_index in range(2, self.worksheet.max_row + 1):
            scenario_dict = {}
            for col_index, header in enumerate(header_row, start=1):
                cell_value = self.worksheet.cell(row=row_index, column=col_index).value
                scenario_dict[header] = cell_value if cell_value is not None else ""

            scenarios_collection.append(scenario_dict)

        print(f" Extracted {len(scenarios_collection)} test scenarios from Excel")
        return scenarios_collection

    def write_test_results(self, results: List[Dict]):
        """
        Writes test results back to Excel file

        Args:
            results: List of dictionaries containing test results
        """
        # Add new columns for results if they don't exist
        header_row = [cell.value for cell in self.worksheet[1]]

        # Check if result columns already exist
        if 'ActualOutcome' not in header_row:
            next_col = len(header_row) + 1
            self.worksheet.cell(row=1, column=next_col).value = 'ActualOutcome'
            self.worksheet.cell(row=1, column=next_col + 1).value = 'ActualMessage'
            self.worksheet.cell(row=1, column=next_col + 2).value = 'TestResult'
            self.worksheet.cell(row=1, column=next_col + 3).value = 'ExecutionTime'

            # Style the new headers
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col in range(next_col, next_col + 4):
                cell = self.worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write results for each test case
        for row_index in range(2, self.worksheet.max_row + 1):
            test_id = self.worksheet.cell(row=row_index, column=1).value

            # Find matching result
            for result in results:
                if result['test_id'] == test_id:
                    # Find the result columns
                    actual_outcome_col = header_row.index(
                        'ActualOutcome') + 1 if 'ActualOutcome' in header_row else len(header_row) + 1

                    # Write actual outcome
                    self.worksheet.cell(row=row_index, column=actual_outcome_col).value = result['actual_status']

                    # Write actual message
                    self.worksheet.cell(row=row_index, column=actual_outcome_col + 1).value = result['actual_message']

                    # Write test result (PASSED/FAILED)
                    result_cell = self.worksheet.cell(row=row_index, column=actual_outcome_col + 2)
                    result_cell.value = result['test_result']

                    # Color code the result
                    from openpyxl.styles import PatternFill
                    if result['test_result'] == 'PASSED':
                        result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                    # Write execution time
                    self.worksheet.cell(row=row_index, column=actual_outcome_col + 3).value = result['execution_time']

                    break

        # Save the workbook
        self.workbook.save(self.filepath)
        print(f"\n Test results written back to Excel file: {self.filepath}")

    def close_connection(self):
        """Closes Excel workbook"""
        if self.workbook:
            self.workbook.close()


class LoginFormValidator:
    """Simulates login validation logic (replaces actual Selenium interaction)"""

    VALID_CREDENTIALS = {
        "student": "Password123"
    }

    @staticmethod
    def validate_login_attempt(username: str, password: str) -> Tuple[str, str]:
        """
        Simulates login validation for practicetestautomation.com
        Valid credentials: username="student", password="Password123"
        Returns: (status, message) tuple
        """

        # Check for empty fields - username is checked first
        if not username:
            return ("FAILURE", "Your username is invalid!")
        if not password:
            return ("FAILURE", "Your password is invalid!")

        # Check for invalid characters (security checks)
        dangerous_patterns = ["<script>", "OR '1'='1", "';", "DROP TABLE", "<", ">"]
        for pattern in dangerous_patterns:
            if pattern.lower() in username.lower():
                return ("FAILURE", "Your username is invalid!")
            if pattern.lower() in password.lower():
                return ("FAILURE", "Your password is invalid!")

        # Check username length (boundary test)
        if len(username) > 100:
            return ("FAILURE", "Your username is invalid!")
        if len(password) > 100:
            return ("FAILURE", "Your password is invalid!")

        # Check for trailing/leading spaces
        if username != username.strip():
            return ("FAILURE", "Your username is invalid!")

        # Validate credentials (case-sensitive)
        if username in LoginFormValidator.VALID_CREDENTIALS:
            if LoginFormValidator.VALID_CREDENTIALS[username] == password:
                return ("SUCCESS", "Logged In Successfully")
            else:
                return ("FAILURE", "Your password is invalid!")
        else:
            return ("FAILURE", "Your username is invalid!")

    @staticmethod
    def perform_login_test(username: str, password: str) -> Dict:
        """
        Executes login test with given credentials
        Returns detailed result dictionary
        """
        print(f"\n Testing credentials: [{username}] / [{'*' * len(password)}]")

        # Simulate page load delay
        time.sleep(0.1)

        # Execute validation
        status, message = LoginFormValidator.validate_login_attempt(username, password)

        result = {
            'actual_status': status,
            'actual_message': message,
            'execution_time': datetime.now().strftime("%H:%M:%S")
        }

        return result


class DataDrivenTestEngine:
    """Main test execution engine"""

    def __init__(self, excel_file: str, sheet_name: str):
        self.data_provider = ExcelDataProvider(excel_file, sheet_name)
        self.tracker = TestExecutionTracker()

    def execute_test_suite(self):
        """Runs complete test suite"""

        print("\n" + "=" * 70)
        print(" DATA-DRIVEN TEST AUTOMATION - LOGIN FUNCTIONALITY")
        print("=" * 70)
        print(f" Target: https://practicetestautomation.com/practice-test-login/")
        print(f" Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 70 + "\n")

        # Initialize Excel connection
        if not self.data_provider.initialize_connection():
            print("Failed to initialize data provider. Exiting...")
            return

        # Extract test scenarios
        test_scenarios = self.data_provider.extract_test_scenarios()

        if not test_scenarios:
            print("No test scenarios found in Excel file.")
            return

        # Execute each test scenario and collect results
        print(f"\n{'─' * 70}")
        print("EXECUTING TEST SCENARIOS")
        print(f"{'─' * 70}")

        results_for_excel = []
        for index, scenario in enumerate(test_scenarios, start=1):
            result_data = self._run_single_test(index, scenario)
            results_for_excel.append(result_data)

        # Generate results
        self._display_detailed_results()

        # Write results back to Excel
        self.data_provider.write_test_results(results_for_excel)

        # Cleanup
        self.data_provider.close_connection()
        print("\n✓ Test execution completed\n")

    def _run_single_test(self, index: int, scenario: Dict) -> Dict:
        """
        Executes individual test case
        """

        test_id = scenario.get('TestCaseID', f'TC{index:03d}')
        description = scenario.get('ScenarioDescription', 'N/A')
        username = scenario.get('InputUsername', '')
        password = scenario.get('InputPassword', '')
        expected_status = scenario.get('ExpectedOutcome', '')
        expected_msg = scenario.get('ExpectedMessage', '')
        category = scenario.get('TestCategory', 'General')

        print(f"\n[{index}] {test_id} - {description}")
        print(f"  Category: {category}")

        # Execute the login test
        test_result = LoginFormValidator.perform_login_test(username, password)

        # Perform assertions
        status_match = test_result['actual_status'] == expected_status
        message_match = expected_msg.lower() in test_result['actual_message'].lower()

        overall_result = "PASSED" if (status_match and message_match) else "FAILED"

        # Log detailed results
        print(f"  Expected: {expected_status} - '{expected_msg}'")
        print(f"  Actual: {test_result['actual_status']} - '{test_result['actual_message']}'")

        if overall_result == "PASSED":
            print(f"  ✓ Result: {overall_result}")
            result_details = f"All assertions passed | {test_result['execution_time']}"
        else:
            print(f"  ✗ Result: {overall_result}")
            failure_reasons = []
            if not status_match:
                failure_reasons.append(
                    f"Status mismatch (Expected: {expected_status}, Got: {test_result['actual_status']})")
            if not message_match:
                failure_reasons.append(
                    f"Message mismatch (Expected: '{expected_msg}', Got: '{test_result['actual_message']}')")
            result_details = " | ".join(failure_reasons)

        # Record in tracker
        self.tracker.record_outcome(test_id, overall_result, result_details)

        # Return result data for Excel logging
        return {
            'test_id': test_id,
            'actual_status': test_result['actual_status'],
            'actual_message': test_result['actual_message'],
            'test_result': overall_result,
            'execution_time': test_result['execution_time']
        }

    def _display_detailed_results(self):
        """Displays comprehensive test results"""

        print("\n" + "=" * 70)
        print(self.tracker.generate_summary_report())

        print("\nDETAILED RESULTS BY TEST CASE:")
        print("─" * 70)

        for outcome in self.tracker.test_outcomes:
            status_symbol = "✓" if outcome['status'] == "PASSED" else "✗"
            status_display = f"{status_symbol} {outcome['status']}"
            print(f"{outcome['id']:<10} {status_display:<15} {outcome['details']}")

        print("=" * 70)


def main():
    """Main entry point"""

    EXCEL_FILENAME = "test_data.xlsx"
    SHEET_NAME = "LoginTestScenarios"

    # Create and run test engine
    test_engine = DataDrivenTestEngine(EXCEL_FILENAME, SHEET_NAME)
    test_engine.execute_test_suite()


if __name__ == "__main__":
    main()