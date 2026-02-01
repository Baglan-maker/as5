#!/usr/bin/env python3
"""
Cross-Browser Testing with BrowserStack
Runs login tests on multiple browsers in the cloud
Compatible with Selenium 4.x
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from datetime import datetime
import time
import config  # Import our configuration file


class BrowserStackTestRunner:
    """Executes tests on BrowserStack cloud platform"""

    def __init__(self):
        self.results = []
        self.driver = None

    def create_driver(self, browser_config):
        """
        Creates a remote WebDriver connected to BrowserStack
        Compatible with Selenium 4.x using options instead of desired_capabilities

        Args:
            browser_config: Dictionary with browser capabilities
        """
        print(f"\n{'=' * 70}")
        print(f"Starting test on: {browser_config['sessionName']}")
        print(f"{'=' * 70}")

        # Determine which browser to use and create appropriate options
        browser_name = browser_config['browserName'].lower()

        if browser_name == 'chrome':
            from selenium.webdriver.chrome.options import Options
            options = Options()
        elif browser_name == 'firefox':
            from selenium.webdriver.firefox.options import Options
            options = Options()
        elif browser_name == 'edge':
            from selenium.webdriver.edge.options import Options
            options = Options()
        elif browser_name == 'safari':
            from selenium.webdriver.safari.options import Options
            options = Options()
        else:
            # Default to Chrome
            from selenium.webdriver.chrome.options import Options
            options = Options()

        # Set browser-specific capabilities
        options.browser_version = browser_config.get('browserVersion', 'latest')
        options.platform_name = browser_config['os']

        # Add BrowserStack specific options
        bstack_options = {
            'os': browser_config['os'],
            'osVersion': browser_config['osVersion'],
            'sessionName': browser_config['sessionName'],
            'buildName': browser_config['buildName'],
            'local': False,
            'seleniumVersion': '4.0.0',
            'debug': True,
            'video': True,
            'networkLogs': True,
            'consoleLogs': 'info'
        }

        options.set_capability('bstack:options', bstack_options)

        # Create remote driver
        self.driver = webdriver.Remote(
            command_executor=config.BS_HUB_URL,
            options=options
        )

        self.driver.implicitly_wait(10)
        return self.driver

    def run_login_test(self, username, password, expected_outcome):
        """
        Executes a single login test

        Args:
            username: Username to test
            password: Password to test
            expected_outcome: Expected result (SUCCESS or FAILURE)

        Returns:
            Dictionary with test results
        """
        try:
            # Navigate to login page
            print(f"  → Navigating to login page...")
            self.driver.get("https://practicetestautomation.com/practice-test-login/")

            # Wait for page to load
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "username"))
            )

            print(f"  → Testing credentials: [{username}] / [{'*' * len(password)}]")

            # Find and fill username field
            username_field = self.driver.find_element(By.ID, "username")
            username_field.clear()
            username_field.send_keys(username)

            # Find and fill password field
            password_field = self.driver.find_element(By.ID, "password")
            password_field.clear()
            password_field.send_keys(password)

            # Click submit button
            submit_button = self.driver.find_element(By.ID, "submit")
            submit_button.click()

            # Wait for response
            time.sleep(2)

            # Check for success or failure
            actual_outcome = "FAILURE"
            actual_message = ""

            try:
                # Check for success message
                success_element = self.driver.find_element(By.CSS_SELECTOR, ".post-title")
                if "Logged In Successfully" in success_element.text or "successfully" in success_element.text.lower():
                    actual_outcome = "SUCCESS"
                    actual_message = "Logged In Successfully"
                    print(f"  ✓ Login successful")
            except:
                # Check for error message
                try:
                    error_element = self.driver.find_element(By.ID, "error")
                    actual_message = error_element.text
                    actual_outcome = "FAILURE"
                    print(f"  ✗ Login failed: {actual_message}")
                except:
                    actual_message = "Unknown response"
                    print(f"  ? Unknown response from application")

            # Determine if test passed
            test_passed = (actual_outcome == expected_outcome)

            result = {
                'username': username,
                'expected': expected_outcome,
                'actual': actual_outcome,
                'message': actual_message,
                'passed': test_passed,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

            if test_passed:
                print(f"  ✓ Test PASSED")
            else:
                print(f"  ✗ Test FAILED (Expected: {expected_outcome}, Got: {actual_outcome})")

            return result

        except Exception as e:
            print(f"  ✗ Error during test: {str(e)}")
            return {
                'username': username,
                'expected': expected_outcome,
                'actual': 'ERROR',
                'message': str(e),
                'passed': False,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

    def execute_test_suite_on_browser(self, browser_config, test_scenarios):
        """
        Runs all test scenarios on a specific browser

        Args:
            browser_config: Browser configuration dictionary
            test_scenarios: List of test scenarios from Excel
        """
        browser_results = []

        try:
            # Create driver for this browser
            self.create_driver(browser_config)

            # Run each test scenario (first 3 tests for demo)
            num_tests = min(3, len(test_scenarios))
            for i, scenario in enumerate(test_scenarios[:num_tests], 1):
                print(f"\n  Test {i}/{num_tests}: {scenario['ScenarioDescription']}")

                result = self.run_login_test(
                    username=scenario['InputUsername'],
                    password=scenario['InputPassword'],
                    expected_outcome=scenario['ExpectedOutcome']
                )

                result['browser'] = browser_config['name']
                result['test_id'] = scenario['TestCaseID']
                browser_results.append(result)

                time.sleep(1)  # Small delay between tests

            # Mark test as passed in BrowserStack
            self.driver.execute_script(
                'browserstack_executor: {"action": "setSessionStatus", '
                '"arguments": {"status":"passed", "reason": "All tests completed"}}'
            )

        except Exception as e:
            print(f"\n✗ Error during test execution: {str(e)}")
            if self.driver:
                try:
                    self.driver.execute_script(
                        'browserstack_executor: {"action": "setSessionStatus", '
                        '"arguments": {"status":"failed", "reason": "' + str(e).replace('"', "'") + '"}}'
                    )
                except:
                    pass
        finally:
            if self.driver:
                print(f"\n  Closing browser...")
                self.driver.quit()

        return browser_results

    def run_all_tests(self):
        """Main method to execute tests on all configured browsers"""

        print("\n" + "=" * 70)
        print("CROSS-BROWSER TESTING WITH BROWSERSTACK")
        print("=" * 70)
        print(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Target: https://practicetestautomation.com/practice-test-login/")
        print(f"Browsers: {len(config.BROWSER_CONFIGS)}")
        print("=" * 70)

        # Load test scenarios from Excel
        print("\n📊 Loading test data from Excel...")
        try:
            wb = openpyxl.load_workbook('test_data.xlsx')
            ws = wb['LoginTestScenarios']

            # Read test scenarios
            test_scenarios = []
            headers = [cell.value for cell in ws[1]]

            for row_idx in range(2, ws.max_row + 1):
                scenario = {}
                for col_idx, header in enumerate(headers, 1):
                    scenario[header] = ws.cell(row_idx, col_idx).value or ""
                test_scenarios.append(scenario)

            wb.close()
            print(f"✓ Loaded {len(test_scenarios)} test scenarios")

        except Exception as e:
            print(f"✗ Error loading Excel file: {e}")
            return

        # Execute tests on each browser
        all_results = []

        for i, browser_config in enumerate(config.BROWSER_CONFIGS, 1):
            print(f"\n{'#' * 70}")
            print(f"BROWSER {i}/{len(config.BROWSER_CONFIGS)}: {browser_config['name']}")
            print(f"{'#' * 70}")

            browser_results = self.execute_test_suite_on_browser(
                browser_config,
                test_scenarios
            )

            all_results.extend(browser_results)

            print(f"\n✓ Completed testing on {browser_config['name']}")
            print(f"  Tests run: {len(browser_results)}")
            print(f"  Passed: {sum(1 for r in browser_results if r['passed'])}")
            print(f"  Failed: {sum(1 for r in browser_results if not r['passed'])}")

        # Display summary
        self.display_summary(all_results)

        print("\n" + "=" * 70)
        print("🎉 ALL TESTS COMPLETED!")
        print("=" * 70)
        print("\n📋 NEXT STEPS:")
        print("1. Go to BrowserStack dashboard: https://automate.browserstack.com/")
        print("2. View your test sessions")
        print("3. Download screenshots and videos")
        print("4. Download execution logs")
        print("5. Include these in your report")
        print("\n")

    def display_summary(self, results):
        """Display summary of all test results"""

        print("\n" + "=" * 70)
        print("TEST EXECUTION SUMMARY")
        print("=" * 70)

        # Group by browser
        browsers = {}
        for result in results:
            browser = result['browser']
            if browser not in browsers:
                browsers[browser] = []
            browsers[browser].append(result)

        # Display results per browser
        for browser, browser_results in browsers.items():
            passed = sum(1 for r in browser_results if r['passed'])
            failed = len(browser_results) - passed
            pass_rate = (passed / len(browser_results) * 100) if browser_results else 0

            print(f"\n{browser}:")
            print(f"  Total Tests: {len(browser_results)}")
            print(f"  Passed: {passed}")
            print(f"  Failed: {failed}")
            print(f"  Pass Rate: {pass_rate:.1f}%")

            # Show individual results
            for result in browser_results:
                status = "✓ PASS" if result['passed'] else "✗ FAIL"
                print(f"    {result['test_id']}: {status} - {result['username']}")

        # Overall summary
        total_tests = len(results)
        total_passed = sum(1 for r in results if r['passed'])
        total_failed = total_tests - total_passed
        overall_pass_rate = (total_passed / total_tests * 100) if total_tests else 0

        print(f"\n{'─' * 70}")
        print("OVERALL SUMMARY:")
        print(f"  Total Tests Across All Browsers: {total_tests}")
        print(f"  Total Passed: {total_passed}")
        print(f"  Total Failed: {total_failed}")
        print(f"  Overall Pass Rate: {overall_pass_rate:.1f}%")
        print("=" * 70)


def main():
    """Main entry point"""

    # Verify credentials are configured
    if config.BS_USERNAME == "yourname_abc123" or config.BS_ACCESS_KEY == "aBcDeFgHiJkLmNoPqRsT":
        print("\n❌ ERROR: Please configure your BrowserStack credentials in config.py")
        print("\nSteps:")
        print("1. Open config.py")
        print("2. Replace BS_USERNAME with your BrowserStack username")
        print("3. Replace BS_ACCESS_KEY with your BrowserStack access key")
        print("4. Save the file and run again")
        print("\nGet credentials from: https://automate.browserstack.com/")
        return

    # Create and run test runner
    runner = BrowserStackTestRunner()
    runner.run_all_tests()


if __name__ == "__main__":
    main()