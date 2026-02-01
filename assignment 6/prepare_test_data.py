#!/usr/bin/env python3
"""
Test Data Preparation Script - CORRECTED VERSION
Generates Excel file with test scenarios for login functionality
Target: https://practicetestautomation.com/practice-test-login/
Valid credentials: Username: student, Password: Password123
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def build_test_data_workbook():
    """Creates structured Excel workbook with test cases"""

    workbook = Workbook()
    test_sheet = workbook.active
    test_sheet.title = "LoginTestScenarios"

    # Define custom styling for professionalism
    header_style = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column headers with unique naming
    column_headers = [
        "TestCaseID",
        "ScenarioDescription",
        "InputUsername",
        "InputPassword",
        "ExpectedOutcome",
        "ExpectedMessage",
        "TestCategory"
    ]

    # Apply header formatting
    for col_idx, header_text in enumerate(column_headers, start=1):
        cell = test_sheet.cell(row=1, column=col_idx)
        cell.value = header_text
        cell.fill = header_style
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border

    # CORRECTED Test data scenarios - using actual valid credentials
    test_scenarios = [
        # Positive test case - ONLY ONE VALID CREDENTIAL
        ["TC001", "Valid student credentials", "student", "Password123", "SUCCESS", "Logged In Successfully",
         "Positive"],
        ["TC002", "Valid credentials with different case password", "student", "password123", "FAILURE",
         "Your password is invalid", "Negative"],
        ["TC003", "Valid username with trailing space", "student ", "Password123", "FAILURE",
         "Your username is invalid", "Negative"],

        # Negative test cases - invalid credentials
        ["TC004", "Invalid username with valid password", "invaliduser", "Password123", "FAILURE",
         "Your username is invalid", "Negative"],
        ["TC005", "Valid username with wrong password", "student", "wrongpassword", "FAILURE",
         "Your password is invalid", "Negative"],
        ["TC006", "Empty username field", "", "Password123", "FAILURE", "Your username is invalid", "Negative"],
        ["TC007", "Empty password field", "student", "", "FAILURE", "Your password is invalid", "Negative"],
        ["TC008", "Both fields empty", "", "", "FAILURE", "Your username is invalid", "Negative"],

        # Security test cases
        ["TC009", "SQL injection in username", "student' OR '1'='1", "Password123", "FAILURE",
         "Your username is invalid", "Security"],
        ["TC010", "XSS attempt in username", "<script>alert('xss')</script>", "Password123", "FAILURE",
         "Your username is invalid", "Security"],
        ["TC011", "SQL injection in password", "student", "' OR '1'='1", "FAILURE", "Your password is invalid",
         "Security"],

        # Boundary test cases
        ["TC012", "Very long username", "a" * 300, "Password123", "FAILURE", "Your username is invalid", "Boundary"],
        ["TC013", "Very long password", "student", "b" * 300, "FAILURE", "Your password is invalid", "Boundary"],
        ["TC014", "Case sensitive username", "Student", "Password123", "FAILURE", "Your username is invalid",
         "Boundary"],
        ["TC015", "Case sensitive username uppercase", "STUDENT", "Password123", "FAILURE", "Your username is invalid",
         "Boundary"],
    ]

    # Populate data rows
    for row_idx, scenario_data in enumerate(test_scenarios, start=2):
        for col_idx, cell_value in enumerate(scenario_data, start=1):
            cell = test_sheet.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.border = cell_border

            # Highlight expected outcomes with colors
            if col_idx == 5:  # ExpectedOutcome column
                if cell_value == "SUCCESS":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Auto-adjust column widths
    column_widths = [12, 35, 30, 20, 18, 30, 15]
    for idx, width in enumerate(column_widths, start=1):
        test_sheet.column_dimensions[test_sheet.cell(row=1, column=idx).column_letter].width = width

    # Add metadata sheet
    metadata_sheet = workbook.create_sheet("TestMetadata")
    metadata_info = [
        ["Property", "Value"],
        ["Test Suite", "Login Functionality Testing"],
        ["Created Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Target Website", "https://practicetestautomation.com/practice-test-login/"],
        ["Valid Credentials", "Username: student | Password: Password123"],
        ["Total Test Cases", len(test_scenarios)],
        ["Framework", "Selenium WebDriver + openpyxl"],
        ["Data Format", "Excel (.xlsx)"]
    ]

    for row_idx, meta_row in enumerate(metadata_info, start=1):
        for col_idx, value in enumerate(meta_row, start=1):
            cell = metadata_sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:
                cell.font = Font(bold=True)

    metadata_sheet.column_dimensions['A'].width = 20
    metadata_sheet.column_dimensions['B'].width = 60

    # Save workbook
    workbook.save("test_data.xlsx")
    print(f"✓ Test data file created successfully: test_data.xlsx")
    print(f"✓ Total scenarios: {len(test_scenarios)}")
    print(f"✓ Valid credentials: student / Password123")
    print(f"✓ Sheets created: {', '.join(workbook.sheetnames)}")


if __name__ == "__main__":
    build_test_data_workbook()